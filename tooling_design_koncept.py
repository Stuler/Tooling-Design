from os import makedirs
from main_2 import dutinka, nadobka, rameno
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Vytvori podadresare projektu
def vytvor_strom_projektu(seznam):
    cesta = input('Zadaj cestu projektu: ')
    for i in seznam:
        makedirs(cesta + "\\" + i)
    # ! Zapracovat algoritmus pre uz existujuci priecinok
    return(cesta)

# Vygeneruje system oznacenia podla pouzivatela
def start_cislo():
    prve_cislo = int(input("Uved koncove cislo prveho stahovacieho kruzku: "))
    return prve_cislo

def vytvor_oznacenie(d_nad, tlak, tvar_ramena, poc_tahov, poc_cislo, h_nad = "h"):
    stat_nazov = []
    cisla_nar = []
    ozn_nar = []
            
    while True:  
        prompt = input("Zadaj moznost:\n"
                "1 - Vygenerovat oznacenie zo zadanych parametrov\n"
                "2 - Vytvorit oznacenie podla uzivatela\n")
        
        if prompt == "1":
                stat_ozn = ("NA-" + d_nad + "-" + h_nad + "-" + tlak + "-" + tvar_ramena)
        elif prompt == "2":
            stat_ozn = input("Zadaj staticku cast oznacenia naradia: ")
        else:
            print("\nNeplatna moznost! Skus znova!")
            continue
                
        for i in range(poc_tahov):
            stat_nazov.append(stat_ozn)
            cisl_ozn = int(poc_cislo) + i
            cisla_nar.append(str(cisl_ozn))
            ozn_nar.append(stat_nazov[i] + "-" + str(cisla_nar[i]))
            
        return ozn_nar

def rozsah_oznacenia(naradie, poc_cislo, poc_tahov):
    poc_cislo = start_cislo()
    prve_cisla = [poc_cislo]
    for i in naradie:
        if poc_tahov % 10 == 0:
            dalsie_cislo = int(((prve_cisla[-1]) + (poc_tahov - 1)) + 1)
        else:
            dalsie_cislo = int(((prve_cisla[-1]) + (poc_tahov - 1) ) / 10) * 10 + 11
        prve_cisla.append(dalsie_cislo)
    
    return(prve_cisla)

### Stahobvaci krouzky ###
def sekvence_dc(d_ram, d_kom, t_kom, t_lak, n):
    d_avg = d_ram - (d_kom + 2*t_kom + 2*t_lak)
    red_avg = d_avg / n 

    sekv_Dc = []
    
    for i in range(1, n+1):
        Dc = round(d_ram - (i * red_avg),2)
        sekv_Dc.append(Dc)

    return sekv_Dc

def sekvence_ds(rozmery_dc):

    sekv_Ds = []
    
    for i in (rozmery_dc):
        Ds = round(i + 0.15, 2)
        sekv_Ds.append(Ds)

    return (sekv_Ds)

def sekvencia_kruzkov(rozmery_Dc, rozmery_Ds, pocet_tahov,cis):
    
    cisla_tahu = []
            
    for i in range(1, rameno.n+1):
        cisla_tahu.append(i)

    oznacenie = vytvor_oznacenie(str(int(nadobka.d_nad)), str(int(nadobka.tlak)),rameno.tvar_ram, rameno.n, cis, str(int(nadobka.h_nad)))
    stah_krouzek = list(zip(cisla_tahu, oznacenie, rozmery_Ds, rozmery_Dc))
    return(stah_krouzek)

### Chytaky ###
def sekvencia_chytakov(dc_kr, poc_tahov,cis):

    sekv_D = []
    sekv_d = []
    cisla_tahu = []

    rozdiel_tlous = nadobka.t_kom - dutinka.t_dut
    rel_rozdiel_tlous = round(rozdiel_tlous / rameno.n, 3)
    tl_steny = [dutinka.t_dut]
    for i in range(1, rameno.n+1):
        tl_steny_tah = dutinka.t_dut + i*rel_rozdiel_tlous
        tl_steny.append(tl_steny_tah)

    for i in range(1, rameno.n+1):
        cisla_tahu.append(i)
        D_ch = round(dc_kr[i-1] - 2*(nadobka.t_vonk_lak + nadobka.t_vnut_lak) - 2*0.03 - 2*(tl_steny[i-1]), 2)
        d_ch = round(D_ch - 5.5, 2)
        sekv_D.append(D_ch)
        sekv_d.append(d_ch)

    oznacenie = vytvor_oznacenie(str(int(nadobka.d_nad)), str(int(nadobka.tlak)),rameno.tvar_ram, rameno.n, cis, str(int(nadobka.h_nad)))
    chytaky = list(zip(cisla_tahu, oznacenie, sekv_D, sekv_d))
    return(chytaky)
 
# Modul vytvori navrhovy excel
# Funkcie modulu sluzia na vytvorenie tabuliek s rozmermi jednotlivych naradi
# a na vytvorenie tabuliek pre parametricke modelovanie sucasti

#### Vytvori navrhovy subor MS Excel ####
def vytvor_data(path, st_kr, chyt, pocet_tahov):
    wb = Workbook()
    dest = path
    ws1 = wb.active
    ws1.title = "Data"
  
    # Prvy stlpec
    ws1['B2'] = "Stena dutinky"
    ws1['B3'] = "Stena ramena"
    ws1['B4'] = "Stena kominku"

    ws1['B6'] = "Pocet tahu"
    ws1['B7'] = "Redukce"

    ws1['B9'] = "Zmena tloustky steny/tah"

    ws1.merge_cells('B11:B13')
    ws1['B11'] = "Tloustka laku"
    ws1['B14'] = "Vule chytaku"

    ws1['B17'] = "Prumer zacatku ramena"

    # Druhy stlpec
    ws1['C2'] = "tdutinka"
    ws1['C3'] = "trameno"
    ws1['C4'] = "tkominek"

    ws1['C6'] = "n"
    ws1['C7'] = "red"

    ws1['C9'] = "tdut/n"

    ws1['C11'] = "tvonklak"
    ws1['C12'] = "tvnutlak"
    ws1['C13'] = "tlakcelk"
    ws1['C14'] = "v"

    ws1['C16'] = "Dnad"
    ws1['C17'] = "Dram"

    # Treti stlpec
    ws1['D2'] = dutinka.t_dut
    ws1['D3'] = rameno.t_ram
    ws1['D4'] = nadobka.t_kom

    ws1['D6'] = rameno.n

    # Vypocet redukcie
    konc_prum = 25.4 + 2*nadobka.t_kom + 2*(nadobka.t_vonk_lak+nadobka.t_vnut_lak)
    rozdiel_priem = rameno.d_ram - konc_prum
    red_tah = rozdiel_priem / rameno.n
    rel_red = red_tah / rameno.d_ram
    ws1['D7'] = round(rel_red * 100, 2)

    # Vypocet zmeny tloustky na tah
    rozdiel_tlous = nadobka.t_kom - dutinka.t_dut
    rel_rozdiel_tlous = round(rozdiel_tlous / rameno.n, 3)
    ws1['D9'] = rel_rozdiel_tlous

    ws1['D11'] = nadobka.t_vonk_lak
    ws1['D12'] = nadobka.t_vnut_lak
    ws1['D13'] = nadobka.t_vonk_lak + nadobka.t_vnut_lak
    ws1['D14'] = rameno.vule

    ws1['D16'] = nadobka.d_nad
    ws1['D17'] = rameno.d_ram

# Stahovaci krouzky
    ws2 = wb.create_sheet(title="Stahovací kroužky")
    ws2['A2'] = "Stahovací kroužky"
    rows_st_krouzky = [
        ["Tah", "Oznaceni", "Ds", "Dc", "Rc", "R", "Lc", "XRc"]
    ]

    for row in rows_st_krouzky:
        ws2.append(row)

    for i in st_kr:
        ws2.append(i)

    ws2['A'+str(3+rameno.n+1)] = "Chytáky"
    rows_chytaky = [
        ["Tah", "Oznaceni", "D", "d"]
    ]
    for row in rows_chytaky:
        ws2.append(row)

    for i in chyt:
        ws2.append(i)
 
    wb.save(dest)
#### Konec tvorby .xls ####

seznam_naradi = [
    "Stahovaci krouzky", 
    "Chytaky", 
    "Vodici pouzdra", 
    "Navadeci krouzky", 
    "Drzaky chytaku",
    "Sroubove cepy",
    "Trny",
    "Pruziny",
    ]

rozsahy = rozsah_oznacenia(seznam_naradi, start_cislo, rameno.n)

d_c = sekvence_dc(rameno.d_ram, 25.4, nadobka.t_kom, nadobka.t_vnut_lak+nadobka.t_vonk_lak, rameno.n)
d_s = sekvence_ds(d_c)
kruzky = sekvencia_kruzkov(d_c, d_s, rameno.n, rozsahy[0])
chytaky = sekvencia_chytakov(d_c, rameno.n, rozsahy[1])

vytvor_data(vytvor_strom_projektu(seznam_naradi)+"\\ navrh_naradi.xlsx", kruzky, chytaky, rameno.n)