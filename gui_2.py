from os import makedirs
from tkinter import *
from nadobka import Nadobka as nad
from nadobka import Dutinka as dut
from nadobka import Rameno as ram
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#####################
### Tvorba excelu ###

def vytvor_data(path, data, pocet_tahov):
    wb = Workbook()
    dest = path
    ws1 = wb.active
    ws1.title = "Data"
  
    # Prvy stlpec
    ws1['B2'] = "Stena dutinky"
    ws1['B3'] = "Stena ramena"
    ws1['B4'] = "Stena kominku"

    ws1['B6'] = "Pocet tahu"
    ws1['B7'] = "Redukce"

    ws1['B9'] = "Zmena tloustky steny/tah"

    ws1.merge_cells('B11:B13')
    ws1['B11'] = "Tloustka laku"
    ws1['B14'] = "Vule chytaku"

    ws1['B17'] = "Prumer zacatku ramena"

    # Druhy stlpec
    ws1['C2'] = "tdutinka"
    ws1['C3'] = "trameno"
    ws1['C4'] = "tkominek"

    ws1['C6'] = "n"
    ws1['C7'] = "red"

    ws1['C9'] = "tdut/n"

    ws1['C11'] = "tvonklak"
    ws1['C12'] = "tvnutlak"
    ws1['C13'] = "tlakcelk"
    ws1['C14'] = "v"

    ws1['C16'] = "Dnad"
    ws1['C17'] = "Dram"

    # Treti stlpec
    ws1['D3'] = tl_ram_value
    ws1['D4'] = tl_kom_value

    ws1['D6'] = poc_tahu_value

    # Vypocet redukcie
    konc_prum = 25.4 + 2*tl_kom_value + 2*(tl_lak_ven_value+tl_lak_vnit_value)
    rozdiel_priem = prum_ram_value - konc_prum
    red_tah = rozdiel_priem / poc_tahu_value
    rel_red = red_tah / prum_ram_value
    ws1['D7'] = round(rel_red * 100, 2)

    # Vypocet zmeny tloustky na tah
    rozdiel_tlous = tl_kom_value - tl_dut_value
    rel_rozdiel_tlous = round(rozdiel_tlous / poc_tahu_value, 3)
    ws1['D9'] = rel_rozdiel_tlous

    ws1['D11'] = tl_lak_ven_value
    ws1['D12'] = tl_lak_vnit_value
    ws1['D13'] = tl_lak_ven_value + tl_lak_vnit_value
    ws1['D14'] = vule_chytak_value

    ws1['D16'] = prum_nad_value
    ws1['D17'] = prum_ram_value

# Stahovaci krouzky
    ws2 = wb.create_sheet(title="Stahovací kroužky")
    ws2['A2'] = "Stahovací kroužky"
    rows = [
        ["Tah", "Oznaceni", "Ds", "Dc", "Rc", "R", "Lc", "XRc"]
    ]
# Legenda
    for row in rows:
        ws2.append(row)
# Pocet tahov
    #for i in range(1, pocet_tahov+1):
    #    ws2.cell(row = 3+i, column = 1).value = i
# Oznaceni
    for i in data:
        ws2.append(i)
 
    wb.save(dest)

#####################
### Navrh naradia ###

def sekvence_dc(d_ram, d_kom, t_kom, t_lak, n):
    d_avg = d_ram - (d_kom + 2*t_kom + 2*t_lak)
    red_avg = d_avg / n 

    sekv_Dc = []
    
    for i in range(1, n+1):
        Dc = round(d_ram - (i * red_avg),2)
        sekv_Dc.append(Dc)

    return sekv_Dc

def sekvence_ds(rozmery_dc):

    sekv_Ds = []
    
    for i in (rozmery_dc):
        Ds = round(i + 0.15, 2)
        sekv_Ds.append(Ds)

    return (sekv_Ds)

def sekvencia_kruzkov(rozmery_Dc, rozmery_Ds, pocet_tahov):
    
    cisla_tahu = []
            
    for i in range(1, pocet_tahov+1):
        cisla_tahu.append(i)

    oznacenie = vytvor_oznacenie(prum_nad_value, tlak_value, uh_ram_value, poc_tahu_value, vys_nad_value)

    stah_krouzek = list(zip(cisla_tahu, oznacenie, rozmery_Ds, rozmery_Dc))
    return(stah_krouzek)

d_c = sekvence_dc(40.23, 25.4, 0.37, 0.04, 19)
d_s = sekvence_ds(d_c)
kruzky = sekvencia_kruzkov(d_c, d_s, 19)

#vytvor_data("C:\\Python\\Test\\navrh_naradi.xlsx", kruzky, 19)
#print(d_c)


#################################
### Tvorba stromu a oznacenia ###
def vytvor_strom_projektu(seznam):
    cesta = input('Zadaj cestu projektu: ')
    for i in seznam:
        makedirs(cesta + "\\" + i)

# Vygeneruje system oznacenia podla pouzivatela
def vytvor_oznacenie(prum_nad_value, tlak_value, tvar_value, poc_tahu_value, vys_nad_value):
    stat_nazov = []
    cisla_nar = []
    ozn_nar = []
        
    while True:  
        prompt = input("Zadaj moznost:\n"
                "1 - Vygenerovat oznacenie zo zadanych parametrov\n"
                "2 - Vytvorit oznacenie podla uzivatela\n")
        
        if prompt == "1":
                stat_ozn = ("NA-" + d_nad + "-" + h_nad + "-" + tlak + "-" + tvar_ramena)
        elif prompt == "2":
            stat_ozn = input("Zadaj staticku cast oznacenia naradia: ")
        else:
            print("\nNeplatna moznost! Skus znova!")
            continue
        
        poc_cislo = input("Uved cislo prveho stahovacieho kruzku: ")
        for i in range(poc_tahov):
            stat_nazov.append(stat_ozn)
            cisl_ozn = int(poc_cislo) + i
            cisla_nar.append(str(cisl_ozn))
            ozn_nar.append(stat_nazov[i] + "-" + str(cisla_nar[i]))
            
        return ozn_nar

def rozsah_oznacenia(naradie, poc_cislo, poc_tahov):
    prve_cisla = [poc_cislo]
    for i in naradie:
        if poc_tahov % 10 == 0:
            dalsie_cislo = int(((prve_cisla[-1]) + (poc_tahov - 1)) + 1)
        else:
            dalsie_cislo = int(((prve_cisla[-1]) + (poc_tahov - 1) ) / 10) * 10 + 11
        prve_cisla.append(dalsie_cislo)
    
    return(prve_cisla)

  
seznam_naradi = [
    "Stahovaci krouzky", 
    "Chytaky", 
    "Vodici pouzdra", 
    "Navadeci krouzky", 
    "Drzaky chytaku",
    "Sroubove cepy",
    "Trny",
    "Pruziny",
    ]

#pocet_tahov = 9
#start_cislo = 101

#stah_kruzky = vytvor_oznacenie("45", "12", "FTP", pocet_tahov)
#print(stah_kruzky)

#rozsahy = rozsah_oznacenia(seznam_naradi, start_cislo, pocet_tahov)
#print(rozsahy)


##################
### Tvorba GUI ###
window=Tk()
window.title("Návrh nářadí")

# Dutinka
dut1 = Label(window, text = "Dutinka")
dut1.grid(row=0, column=0)

dut2 = Label(window, text = "Průměr dutinky")
dut2.grid(row=1, column=1)

dut3 = Label(window, text = "Tloušťka dutinky")
dut3.grid(row=2, column=1)

dut4 = Label(window, text = "Výška dutinky")
dut4.grid(row=3, column=1)

prum_dut_value = DoubleVar()
prum_dut = Entry(window,textvariable=prum_dut_value, width=5)
prum_dut.grid(row=1, column=2)

tl_dut_value = DoubleVar()
tl_dut = Entry(window,textvariable=tl_dut_value, width=5)
tl_dut.grid(row=2, column=2)

vys_dut_value = IntVar()
vys_dut = Entry(window,textvariable=vys_dut_value, width=5)
vys_dut.grid(row=3, column=2)

jedn1 = Label(window, text = "mm")
jedn1.grid(row=1, column=3)

jedn2 = Label(window, text = "mm")
jedn2.grid(row=2, column=3)

jedn3 = Label(window, text = "mm")
jedn3.grid(row=3, column=3)

# Nadobka
nad1 = Label(window, text = "Nádobka")
nad1.grid(row=4, column=0)

nad2 = Label(window, text = "Průměr nádobky")
nad2.grid(row=5, column=1)

nad3 = Label(window, text = "Výška nádobky")
nad3.grid(row=6, column=1)

nad4 = Label(window, text = "Tloušťka komínku")
nad4.grid(row=7, column=1)

nad5 = Label(window, text = "Tloušťka vnitřního laku")
nad5.grid(row=8, column=1)

nad6 = Label(window, text = "Tloušťka vnějšího laku")
nad6.grid(row=9, column=1)

nad7 = Label(window, text = "Průměr komínku")
nad7.grid(row=10, column=1)

nad8 = Label(window, text = "Tlaková specifikace")
nad8.grid(row=11, column=1)

nad9 = Label(window, text = "Tvar ramena")
nad8.grid(row=12, column=1)

prum_nad_value = DoubleVar()
prum_nad = Entry(window,textvariable=prum_nad_value, width=5)
prum_nad.grid(row=5, column=2)

vys_nad_value = IntVar()
vys_nad = Entry(window,textvariable=vys_nad_value, width=5)
vys_nad.grid(row=6, column=2)

tl_kom_value = DoubleVar()
tl_kom = Entry(window,textvariable=tl_kom_value, width=5)
tl_kom.grid(row=7, column=2)

tl_lak_vnit_value = DoubleVar()
tl_lak_vnit = Entry(window,textvariable=tl_lak_vnit_value, width=5)
tl_lak_vnit.grid(row=8, column=2)

tl_lak_ven_value = DoubleVar()
tl_lak_ven = Entry(window,textvariable=tl_lak_ven_value, width=5)
tl_lak_ven.grid(row=9, column=2)

prum_kom_value = DoubleVar()
prum_kom = Entry(window,textvariable=prum_kom_value, width=5)
prum_kom.grid(row=10, column=2)

tlak_value = IntVar()
tlak = Entry(window,textvariable=tlak_value, width=5)
tlak.grid(row=11, column=2)

tvar_value = StringVar()
tvar = Entry(window,textvariable=tvar_value, width=5)
tvar.grid(row=12, column=2)

jedn4 = Label(window, text = "mm")
jedn4.grid(row=5, column=3)

jedn5 = Label(window, text = "mm")
jedn5.grid(row=6, column=3)

jedn6 = Label(window, text = "mm")
jedn6.grid(row=7, column=3)

jedn7 = Label(window, text = "mm")
jedn7.grid(row=8, column=3)

jedn8 = Label(window, text = "mm")
jedn8.grid(row=9, column=3)

jedn9 = Label(window, text = "mm")
jedn9.grid(row=10, column=3)

jedn10 = Label(window, text = "bar")
jedn10.grid(row=11, column=3)

# rameno
ram1 = Label(window, text = "Rameno")
ram1.grid(row=13, column=0)

ram2 = Label(window, text = "Průměr ramena")
ram2.grid(row=14, column=1)

ram3 = Label(window, text = "Tloušťka ramena")
ram3.grid(row=15, column=1)

ram4 = Label(window, text = "Ůhel ramena")
ram4.grid(row=19, column=1)

ram5 = Label(window, text = "Počet tahu")
ram5.grid(row=17, column=1)

ram6 = Label(window, text = "Vule chytáku")
ram6.grid(row=18, column=1)

prum_ram_value = DoubleVar()
prum_ram = Entry(window,textvariable=prum_ram_value, width=5)
prum_ram.grid(row=14, column=2)

tl_ram_value = DoubleVar()
tl_ram = Entry(window,textvariable=tl_ram_value, width=5)
tl_ram.grid(row=15, column=2)

uh_ram_value = DoubleVar()
uh_ram = Entry(window,textvariable=uh_ram_value, width=5)
uh_ram.grid(row=19, column=2)

poc_tahu_value = IntVar()
poc_tahu = Entry(window,textvariable=poc_tahu_value, width=5)
poc_tahu.grid(row=17, column=2)

vule_chytak_value = DoubleVar()
vule_chytak = Entry(window,textvariable=vule_chytak_value, width=5)
vule_chytak.grid(row=18, column=2)

jedn11 = Label(window, text = "mm")
jedn11.grid(row=14, column=3)

jedn12 = Label(window, text = "mm")
jedn12.grid(row=15, column=3)

jedn13 = Label(window, text = "°")
jedn13.grid(row=19, column=3)

jedn14 = Label(window, text = "-")
jedn14.grid(row=17, column=3)

jedn15 = Label(window, text = "mm")
jedn15.grid(row=18, column=3)

# Tlacitka
b1=Button(window, text="OK", command=vytvor_data("C:\\Python\\Test\\navrh_naradi.xlsx", kruzky, 19), width = 10)
b1.grid(row=20, column=0)

b2=Button(window, text="Reset", width = 10)
b2.grid(row=20, column=1)

b3=Button(window, text="Konec", width = 10)
b3.grid(row=20, column=2)

window.mainloop()