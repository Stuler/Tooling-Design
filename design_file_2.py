# Modul vytvori navrhovy excel
# Funkcie modulu sluzia na vytvorenie tabuliek s rozmermi jednotlivych naradi
# a na vytvorenie tabuliek pre parametricke modelovanie sucasti

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import gui

# Vytvori navrhovy subor MS Excel
def vytvor_data(path, data, pocet_tahov):
    wb = Workbook()
    dest = path
    ws1 = wb.active
    ws1.title = "Data"
  
    # Prvy stlpec
    ws1['B2'] = "Stena dutinky"
    ws1['B3'] = "Stena ramena"
    ws1['B4'] = "Stena kominku"

    ws1['B6'] = "Pocet tahu"
    ws1['B7'] = "Redukce"

    ws1['B9'] = "Zmena tloustky steny/tah"

    ws1.merge_cells('B11:B13')
    ws1['B11'] = "Tloustka laku"
    ws1['B14'] = "Vule chytaku"

    ws1['B17'] = "Prumer zacatku ramena"

    # Druhy stlpec
    ws1['C2'] = "tdutinka"
    ws1['C3'] = "trameno"
    ws1['C4'] = "tkominek"

    ws1['C6'] = "n"
    ws1['C7'] = "red"

    ws1['C9'] = "tdut/n"

    ws1['C11'] = "tvonklak"
    ws1['C12'] = "tvnutlak"
    ws1['C13'] = "tlakcelk"
    ws1['C14'] = "v"

    ws1['C16'] = "Dnad"
    ws1['C17'] = "Dram"

    # Treti stlpec
    ws1['D3'] = gui.tl_ram_value
    ws1['D4'] = gui.tl_kom_value

    ws1['D6'] = gui.poc_tahu_value

    # Vypocet redukcie
    konc_prum = 25.4 + 2*gui.tl_kom_value + 2*(gui.tl_lak_ven_value+gui.tl_lak_vnit_value)
    rozdiel_priem = gui.prum_ram_value - konc_prum
    red_tah = rozdiel_priem / gui.poc_tahu_value
    rel_red = red_tah / gui.prum_ram_value
    ws1['D7'] = round(rel_red * 100, 2)

    # Vypocet zmeny tloustky na tah
    rozdiel_tlous = gui.tl_kom_value - gui.tl_dut_value
    rel_rozdiel_tlous = round(rozdiel_tlous / gui.poc_tahu_value, 3)
    ws1['D9'] = rel_rozdiel_tlous

    ws1['D11'] = gui.tl_lak_ven_value
    ws1['D12'] = gui.tl_lak_vnit_value
    ws1['D13'] = gui.tl_lak_ven_value + gui.tl_lak_vnit_value
    ws1['D14'] = gui.vule_chytak_value

    ws1['D16'] = gui.prum_nad_value
    ws1['D17'] = gui.prum_ram_value

# Stahovaci krouzky
    ws2 = wb.create_sheet(title="Stahovací kroužky")
    ws2['A2'] = "Stahovací kroužky"
    rows = [
        ["Tah", "Oznaceni", "Ds", "Dc", "Rc", "R", "Lc", "XRc"]
    ]
# Legenda
    for row in rows:
        ws2.append(row)
# Pocet tahov
    #for i in range(1, pocet_tahov+1):
    #    ws2.cell(row = 3+i, column = 1).value = i
# Oznaceni
    for i in data:
        ws2.append(i)
 
    wb.save(dest)