import os
from inicializacia_projektu import vytvor_oznacenie as ozn
from inicializacia_projektu import vytvor_strom_projektu as strom
from main import dutinka, nadobka, rameno
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def sekvence_dc(d_ram, d_kom, t_kom, t_lak, n):
    d_avg = d_ram - (d_kom + 2*t_kom + 2*t_lak)
    red_avg = d_avg / n 

    sekv_Dc = []
    
    for i in range(1, n+1):
        Dc = round(d_ram - (i * red_avg),2)
        sekv_Dc.append(Dc)

    return sekv_Dc

def sekvence_ds(rozmery_dc):

    sekv_Ds = []
    
    for i in (rozmery_dc):
        Ds = round(i + 0.15, 2)
        sekv_Ds.append(Ds)

    return (sekv_Ds)

def sekvencia_kruzkov(rozmery_Dc, rozmery_Ds, pocet_tahov):
    
    cisla_tahu = []
            
    for i in range(1, pocet_tahov+1):
        cisla_tahu.append(i)

    
    oznacenie = ozn(str(int(nadobka.d_nad)), str(int(nadobka.tlak)),rameno.tvar_ram, 19, str(int(nadobka.h_nad)))
    stah_krouzek = list(zip(cisla_tahu, oznacenie, rozmery_Ds, rozmery_Dc))
    return(stah_krouzek)


# Modul vytvori navrhovy excel
# Funkcie modulu sluzia na vytvorenie tabuliek s rozmermi jednotlivych naradi
# a na vytvorenie tabuliek pre parametricke modelovanie sucasti

# Vytvori navrhovy subor MS Excel
def vytvor_data(path, data, pocet_tahov):
    wb = Workbook()
    dest = path
    ws1 = wb.active
    ws1.title = "Data"
  
    # Prvy stlpec
    ws1['B2'] = "Stena dutinky"
    ws1['B3'] = "Stena ramena"
    ws1['B4'] = "Stena kominku"

    ws1['B6'] = "Pocet tahu"
    ws1['B7'] = "Redukce"

    ws1['B9'] = "Zmena tloustky steny/tah"

    ws1.merge_cells('B11:B13')
    ws1['B11'] = "Tloustka laku"
    ws1['B14'] = "Vule chytaku"

    ws1['B17'] = "Prumer zacatku ramena"

    # Druhy stlpec
    ws1['C2'] = "tdutinka"
    ws1['C3'] = "trameno"
    ws1['C4'] = "tkominek"

    ws1['C6'] = "n"
    ws1['C7'] = "red"

    ws1['C9'] = "tdut/n"

    ws1['C11'] = "tvonklak"
    ws1['C12'] = "tvnutlak"
    ws1['C13'] = "tlakcelk"
    ws1['C14'] = "v"

    ws1['C16'] = "Dnad"
    ws1['C17'] = "Dram"

    # Treti stlpec
    ws1['D3'] = rameno.t_ram
    ws1['D4'] = nadobka.t_kom

    ws1['D6'] = rameno.n

    # Vypocet redukcie
    konc_prum = 25.4 + 2*nadobka.t_kom + 2*(nadobka.t_vonk_lak+nadobka.t_vnut_lak)
    rozdiel_priem = rameno.d_ram - konc_prum
    red_tah = rozdiel_priem / rameno.n
    rel_red = red_tah / rameno.d_ram
    ws1['D7'] = round(rel_red * 100, 2)

    # Vypocet zmeny tloustky na tah
    rozdiel_tlous = nadobka.t_kom - dutinka.t_dut
    rel_rozdiel_tlous = round(rozdiel_tlous / rameno.n, 3)
    ws1['D9'] = rel_rozdiel_tlous

    ws1['D11'] = nadobka.t_vonk_lak
    ws1['D12'] = nadobka.t_vnut_lak
    ws1['D13'] = nadobka.t_vonk_lak + nadobka.t_vnut_lak
    ws1['D14'] = rameno.n

    ws1['D16'] = nadobka.d_nad
    ws1['D17'] = rameno.d_ram


# Stahovaci krouzky
    ws2 = wb.create_sheet(title="Stahovací kroužky")
    ws2['A2'] = "Stahovací kroužky"
    rows = [
        ["Tah", "Oznaceni", "Ds", "Dc", "Rc", "R", "Lc", "XRc"]
    ]
# Legenda
    for row in rows:
        ws2.append(row)
# Pocet tahov
    #for i in range(1, pocet_tahov+1):
    #    ws2.cell(row = 3+i, column = 1).value = i
# Oznaceni
    
    for i in data:
        ws2.append(i)
 
    wb.save(dest)

seznam_naradi = [
    "Stahovaci krouzky", 
    "Chytaky", 
    "Vodici pouzdra", 
    "Navadeci krouzky", 
    "Drzaky chytaku",
    "Sroubove cepy",
    "Trny",
    "Pruziny",
    ]

d_c = sekvence_dc(rameno.d_ram, 25.4, nadobka.t_kom, nadobka.t_vnut_lak+nadobka.t_vonk_lak, rameno.n)
d_s = sekvence_ds(d_c)
kruzky = sekvencia_kruzkov(d_c, d_s, 19)

vytvor_data(strom(seznam_naradi)+"\\ navrh_naradi.xlsx", kruzky, rameno.n)
#print(d_c)