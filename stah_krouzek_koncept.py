from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def sekvencia_kruzkov(d_ram, d_kom, t_kom, t_lak, n):
    
    d_avg = d_ram - (d_kom + 2*t_kom + 2*t_lak)
    red_avg = d_avg / n 

    cisla_tahu = []
    sekv_Dc = []
    sekv_Ds = []
        
    for i in range(1, n+1):
        cisla_tahu.append(i)

    for i in range(1, n+1):
        Dc = round(d_ram - (i * red_avg),2)
        sekv_Dc.append(Dc)

    for i in (sekv_Dc):
        Ds = round(i + 0.15, 2)
        sekv_Ds.append(Ds)

    stah_krouzek = list(zip(cisla_tahu, sekv_Dc, sekv_Ds))
    return(stah_krouzek)

# Vytvori navrhovy subor MS Excel
def vytvor_xls(cesta, data, pocet_tahov):
    wb = Workbook()
    dest = cesta + "\\navrh_naradi.xlsx"
    ws = wb.active
    ws.title = "Naradi"
    
    for row in range(1, pocet_tahov + 1):
        for col in range(1, 3):
            _ = ws.cell(row = row, column = col, value = sekvencia_kruzkov(40.23, 25.4, 0.37, 0.04, 19))

    wb.save(dest)

print(sekvencia_kruzkov(40.23, 25.4, 0.37, 0.04, 19))
vytvor_xls('C:\\Python', sekvencia_kruzkov, 19)