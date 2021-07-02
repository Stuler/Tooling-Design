from os import makedirs, walk, listdir
import os.path
from shutil import copy2, copytree
from tkinter import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

#TODO: REFACTOR!!!
#TODO: Zaskrtat vsetky naradia

class Window():
    def __init__(self, window):
        self.window = window
        self.window.wm_title("Návrh nářadí")

        # Nadobka
        Label(window, text = "Nádobka").grid(row=0, column=0, sticky=W)
        Label(window, text = "Průměr nádobky").grid(row=1, column=1, sticky=W)
        Label(window, text = "Výška nádobky").grid(row=2, column=1, sticky=W)
        Label(window, text = "Tloušťka komínku").grid(row=3, column=1, sticky=W)
        Label(window, text = "Tloušťka vnitřního laku").grid(row=4, column=1, sticky=W)
        Label(window, text = "Tloušťka vnějšího laku").grid(row=5, column=1, sticky=W)
        Label(window, text = "Průměr komínku").grid(row=6, column=1, sticky=W)
        Label(window, text = "Tlaková specifikace").grid(row=7, column=1, sticky=W)

        self.prumNad_value = IntVar(value=45)
        self.prumNad = Entry(window,textvariable=self.prumNad_value, width=5)
        self.prumNad.grid(row=1, column=2)

        self.vysNad_value = IntVar(value=159)
        self.vysNad = Entry(window,textvariable=self.vysNad_value, width=5)
        self.vysNad.grid(row=2, column=2)

        self.tlKom_value = DoubleVar(value=0.37)
        self.tlKom = Entry(window,textvariable=self.tlKom_value, width=5)
        self.tlKom.grid(row=3, column=2)

        self.tlIntLak_value = DoubleVar(value=0.02)
        self.tlIntLak = Entry(window,textvariable=self.tlIntLak_value, width=5)
        self.tlIntLak.grid(row=4, column=2)

        self.tlExtLak_value = DoubleVar(value=0.02)
        self.tlExtLak = Entry(window,textvariable=self.tlExtLak_value, width=5)
        self.tlExtLak.grid(row=5, column=2)

        self.prumKom_value = DoubleVar(value=25.4)
        self.prumKom = Entry(window,textvariable=self.prumKom_value, width=5)
        self.prumKom.grid(row=6, column=2)

        self.tlak_value = IntVar(value=12)
        self.tlak = Entry(window,textvariable=self.tlak_value, width=5)
        self.tlak.grid(row=7, column=2)
        
        Label(window, text = "mm").grid(row=1, column=3, sticky=W)
        Label(window, text = "mm").grid(row=2, column=3, sticky=W)
        Label(window, text = "mm").grid(row=3, column=3, sticky=W)
        Label(window, text = "mm").grid(row=4, column=3, sticky=W)
        Label(window, text = "mm").grid(row=5, column=3, sticky=W)
        Label(window, text = "mm").grid(row=6, column=3, sticky=W)
        Label(window, text = "bar").grid(row=7, column=3, sticky=W)

        # rameno
        Label(window, text = "Rameno").grid(row=8, column=0, sticky=W)
        Label(window, text = "Průměr ramena").grid(row=9, column=1, sticky=W)
        Label(window, text = "Tloušťka ramena").grid(row=10, column=1, sticky=W)
        Label(window, text = "Vule chytáku").grid(row=11, column=1, sticky=W)
        Label(window, text = "Ůhel ramena").grid(row=12, column=1, sticky=W)
        Label(window, text = "Počet tahů").grid(row=13, column=1, sticky=W)
        Label(window, text = "Tvar ramene").grid(row=14, column=1, sticky=W)
        Label(window, text = "Redukce").grid(row=15, column=1, sticky=W)

        self.prumRam_value = DoubleVar(value=40.23)
        self.prumRam = Entry(window,textvariable=self.prumRam_value, width=5)
        self.prumRam.grid(row=9, column=2)

        self.tlRam_value = DoubleVar(value=0.24)
        self.tlRam = Entry(window,textvariable=self.tlRam_value, width=5)
        self.tlRam.grid(row=10, column=2)

        self.vuleChyt_value = DoubleVar(value=0.03)
        self.vuleChyt = Entry(window,textvariable=self.vuleChyt_value, width=5)
        self.vuleChyt.grid(row=11, column=2)

        self.uhRam_value = DoubleVar(value=27)
        self.uhRam = Entry(window,textvariable=self.uhRam_value, width=5)
        self.uhRam.grid(row=12, column=2)

        self.pocTah_value = IntVar(value=19)
        self.pocTah = Entry(window,textvariable=self.pocTah_value, width=5)
        self.pocTah.grid(row=13, column=2)

        self.tvarRam_value = StringVar(value="Crown")
        self.tvarRam = Entry(window,textvariable=self.tvarRam_value, width=5)
        self.tvarRam.grid(row=14, column=2)

        self.prumRed_value = DoubleVar()
        self.prumRed = Entry(window,textvariable=self.prumRed_value, width=5)
        self.prumRed.grid(row=15, column=2)

        Label(window, text = "mm").grid(row=9, column=3, sticky=W)
        Label(window, text = "mm").grid(row=10, column=3, sticky=W)
        Label(window, text = "mm").grid(row=11, column=3, sticky=W)
        Label(window, text = "°").grid(row=12, column=3, sticky=W)
        Label(window, text = "-").grid(row=13, column=3, sticky=W)
        Label(window, text = "-").grid(row=14, column=3, sticky=W)
        Label(window, text = "%").grid(row=15, column=3, sticky=W)
        
        # Seznam naradi
        Label(window, text = "Seznam nářadí").grid(row=0, column=5, sticky=W)

        self.stKr = IntVar()
        Checkbutton(window, text="Stahovací kroužky", variable=self.stKr).grid(row=1, column=5, sticky=W)

        self.chytaky = IntVar()
        Checkbutton(window, text="Chytáky", variable=self.chytaky).grid(row=2, column=5, sticky=W)

        self.vodPzdra = IntVar()
        Checkbutton(window, text="Vodící pouzdra", variable=self.vodPzdra).grid(row=3, column=5, sticky=W)

        self.navKr = IntVar()
        Checkbutton(window, text="Naváděcí kroužky", variable=self.navKr).grid(row=4, column=5, sticky=W)

        self.drzChyt = IntVar()
        Checkbutton(window, text="Držáky chytáků", variable=self.drzChyt).grid(row=5, column=5, sticky=W)

        self.sroubCepy = IntVar()
        Checkbutton(window, text="Šroubové čepy", variable=self.sroubCepy).grid(row=6, column=5, sticky=W)

        self.trny = IntVar()
        Checkbutton(window, text="Trny", variable=self.trny).grid(row=7, column=5, sticky=W)

        self.pruziny = IntVar()
        Checkbutton(window, text="Pružiny", variable=self.pruziny).grid(row=8, column=5, sticky=W)
        Label(window, text = "Startovací číslo").grid(row=0, column=6, sticky=W)

        # Startovacie cisla
        self.stKrInitVal = IntVar(value=101)
        self.stKrInit = Entry(window,textvariable=self.stKrInitVal, width=5)
        self.stKrInit.grid(row=1, column=6)

        self.chytakyInitVal = IntVar(value=201)
        self.chytakyInit = Entry(window,textvariable=self.chytakyInitVal, width=5)
        self.chytakyInit.grid(row=2, column=6)

        self.vodPzdraInitVal = IntVar(value=301)
        self.vodPzdraInit = Entry(window,textvariable=self.vodPzdraInitVal, width=5)
        self.vodPzdraInit.grid(row=3, column=6)

        self.navKrInitVal = IntVar(value=401)
        self.navKrInit = Entry(window,textvariable=self.navKrInitVal, width=5)
        self.navKrInit.grid(row=4, column=6)

        self.drzChytInitVal = IntVar(value=501)
        self.drzChytInit = Entry(window,textvariable=self.drzChytInitVal, width=5)
        self.drzChytInit.grid(row=5, column=6)

        self.sroubCepyInitVal = IntVar(value=601)
        self.sroubCepyInit = Entry(window,textvariable=self.sroubCepyInitVal, width=5)
        self.sroubCepyInit.grid(row=6, column=6)

        self.trnyInitVal = IntVar(value=701)
        self.trnyInit = Entry(window,textvariable=self.trnyInitVal, width=5)
        self.trnyInit.grid(row=7, column=6)

        self.pruzinyInitVal = IntVar(value=801)
        self.pruzinyInit = Entry(window,textvariable=self.pruzinyInitVal, width=5)
        self.pruzinyInit.grid(row=8, column=6)

        Label(window, text = "Číslo prvního tahu").grid(row=1, column=7, sticky=W)
        self.prvniTahVal = IntVar(value=1)
        self.prvniTah = Entry(window,textvariable=self.prvniTahVal, width=2)
        self.prvniTah.grid(row=1, column=8)

        Label(window, text = "Počet tvarových stah. kroužků").grid(row=2, column=7, sticky=W)
        self.pocTvarKrVal = IntVar(value=2)
        self.pocTvarKr = Entry(window,textvariable=self.pocTvarKrVal, width=2)
        self.pocTvarKr.grid(row=2, column=8)

        # Cesta projektu
        projectPath = Label(window, text = "Cesta projektu:")
        projectPath.grid(row=21, column=0, sticky=W)

        self.projPath_value = StringVar(value= "C:\Python\Test")
        self.projPath = Entry(window,textvariable=self.projPath_value, width=40)
        self.projPath.grid(row=21, column=1, columnspan=3)
        
        # Oznacenie naradia
        oznNar = Label(window, text = "Oznacenie naradia:")
        oznNar.grid(row=22, column=0, sticky=W)

        self.oznNar_value = StringVar()
        self.oznNar = Entry(window,textvariable=self.oznNar_value, width=40)
        self.oznNar.grid(row=22, column=1, columnspan=3)

        # Tlacitka
        b1=Button(window, text="OK",  width = 10, command=self.exec)
        b1.grid(row=23, column=0)

        b2=Button(window, text="Reset", width = 10)
        b2.grid(row=23, column=1)

        b3=Button(window, text="Konec", width = 10, command=self.window.destroy)
        b3.grid(row=23, column=2)

        b4=Button(window, text="Vytvor", width = 5, command=self.createDir)
        b4.grid(row=21, column=4)

        b5=Button(window, text="Vytvor", width = 5,command=self.set_numbering)
        b5.grid(row=22, column=4)

        b6=Button(window, text="Přepočítat", width = 8,command=self.set_red)
        b6.grid(row=15, column=4)

        # cesta k sablonam
        sabNar = Label(window, text = "Cesta k šablonám nářadí:")
        sabNar.grid(row=9, column=5, sticky=W)

        self.sabNar_value = StringVar(value = "M:\R&D projekty\Ostatní\Jurek\Sablony_naradia")
        self.sabNar = Entry(window,textvariable=self.sabNar_value, width=40)
        self.sabNar.grid(row=9, column=6, columnspan=3)

#################################################
################## HLAVNY EXCEL #################
#################################################

    def create_XLS(self):
        wb = Workbook()
        dest = self.projPath_value.get() + "\\navrh_naradi.xlsx"
        ws1 = wb.active
        ws1.title = "Data"
        ws2 = wb.create_sheet(title="Rozmery naradi")
        ws1.column_dimensions["B"].width = 25.0
        ws2.column_dimensions["B"].width = 25.0

#################################################
################### Prvy Sheet ##################
#################################################

        # Prvy stlpec
        ws1['B2'] = "Stena dutinky"
        ws1['B3'] = "Stena ramena"
        ws1['B4'] = "Stena kominku"

        ws1['B6'] = "Pocet tahu"
        ws1['B7'] = "Redukce"

        ws1['B9'] = "Zmena tloustky steny/tah"

        ws1.merge_cells('B11:B13')
        ws1['B11'] = "Tloustka laku"
        ws1['B14'] = "Vule chytaku"

        ws1['B16'] = "Prumer nadobky"
        ws1['B17'] = "Prumer zacatku ramena"
        ws1['B18'] = "Prumer kominku"

        # Druhy stlpec
        ws1['C2'] = "t_dut"
        ws1['C3'] = "t_ram"
        ws1['C4'] = "t_kom"

        ws1['C6'] = "n"
        ws1['C7'] = "red"

        ws1['C9'] = "t_dut/n"

        ws1['C11'] = "t_vonklak"
        ws1['C12'] = "t_vnutlak"
        ws1['C13'] = "t_lakcelk"
        ws1['C14'] = "v"

        ws1['C16'] = "D_nad"
        ws1['C17'] = "D_ram"
        ws1['C18'] = "D_kom"

        # Treti stlpec
        ws1['D2'] = float(self.tlRam.get())
        ws1['D3'] = float(self.tlRam.get())
        ws1['D4'] = float(self.tlKom.get())

        ws1['D6'] = int(self.pocTah.get())
        ws1['D7'] = "=(((D17-(D18+2*D4+2*D13))/(D6))/((D17+D18)/2))"

        ws1['D9'] = "=((D4-D3)/D6)"

        ws1['D11'] = float(self.tlIntLak.get())
        ws1['D12'] = float(self.tlExtLak.get())  
        ws1['D13'] = "=D11+D12"
        ws1['D14'] = float(self.vuleChyt.get())

        ws1['D16'] = float(self.prumNad.get())
        ws1['D17'] = float(self.prumRam.get())
        ws1['D18'] = float(self.prumKom.get())

        ws1['D7'].number_format = '0.00 %'
        ws1['D9'].number_format = '0.0000'

        # Stvrty stlpec
        ws1['E2'] = "mm"
        ws1['E3'] = "mm"
        ws1['E4'] = "mm"

        ws1['E6'] = "-"
        ws1['E7'] = "-"

        ws1['E9'] = "mm/tah"

        ws1['E11'] = "mm"
        ws1['E12'] = "mm"  
        ws1['E13'] = "mm"
        ws1['E14'] = "mm"

        ws1['E16'] = "mm"
        ws1['E17'] = "mm"      
        ws1['E18'] = "mm"   

#################################################
#################### Naradie ####################
#################################################
        
        # Stahovaci krouzky
        ws2['A2'] = "Stahovací kroužky"
        rows_st_krouzky = [
                    ["Tah", "Oznaceni", "Ds", "Dc", "Rc", "R", "Lc", "XRc"]
                        ]
        for row in rows_st_krouzky:
            ws2.append(row)

        # Vyplneni tahu
        ws2['A4'] = int(self.prvniTah.get())
        for i in range(5, 5+int(self.pocTah.get())-1):
            ws2['A'+str(i)] ='=A'+str(i-1)+'+1'

        # Vyplneni oznaceni
        ws2['B4'] = self.oznNar.get() + self.stKrInit.get()
        for i in range(5, 5+int(self.pocTah.get())-1):
            koncCis = int(ws2['B'+ str(i-1)].value[-3:])
            ws2['B'+ str(i)] = self.oznNar.get() + str(koncCis+1)

        # Vypocet Dc
        ws2['D4'] = "=Data!D17*(1-DATA!D7)"
        for i in range(5, 5+int(self.pocTah.get())-1):
            ws2['D'+ str(i)] = '=D'+str(i-1)+'*(1-DATA!D7)'
        for i in range(4, 5+int(self.pocTah.get())-1):
            ws2['D'+ str(i)].number_format = '0.00'

        # Vypocet Ds
        for i in range(4, 4+int(self.pocTah.get())):
            ws2['C'+str(i)] = '=D'+str(i)+'+0.15'
            ws2['C'+ str(i)].number_format = '0.00'

        # Chytaky
        pilotsOffset = 4+int(self.pocTah.get())+1
        pilotsLabelRow = pilotsOffset + 1
        pilotsAutoFunctRow = pilotsOffset + 2
        pilotsFunctRow = pilotsOffset + 3

        ws2['A'+str(pilotsOffset)] = "Chytáky"
        rows_st_krouzky = [
                    ["Tah", "Oznaceni", "D", "d"]
                        ]
        for row in rows_st_krouzky:
            ws2.append(row)

        # Vyplneni tahu
        ws2['A'+str(pilotsAutoFunctRow)] = int(self.prvniTah.get())
        for i in range(pilotsFunctRow, pilotsFunctRow+int(self.pocTah.get())-1):
            ws2['A'+str(i)] ='=A'+str(i-1)+'+1'

        # Vyplneni oznaceni
        ws2['B'+str(pilotsAutoFunctRow)] = self.oznNar.get() + self.chytakyInit.get()
        for i in range(pilotsFunctRow, pilotsFunctRow+int(self.pocTah.get())-1):
            koncCis = int(ws2['B'+ str(i-1)].value[-3:])
            ws2['B'+ str(i)] = self.oznNar.get() + str(koncCis+1)       
        
        # Vypocet D
        strokeLst = []
        refLst = []
        for i in range(int(self.pocTah.get())):
            strokeLst.append(i+1)
        for i in strokeLst:
            ws2['C'+str(i+pilotsLabelRow)] = "=D"+str(i+3)+"-2*DATA!$D$13-2*DATA!$D$14-2*(Data!$D$2+"+str(i)+"*Data!$D$9)"
            ws2['C'+str(i+pilotsLabelRow)].number_format = '0.00'

        # Vypocet D
        for i in range(pilotsAutoFunctRow, pilotsAutoFunctRow+int(self.pocTah.get())):
            ws2['D'+str(i)] = '=C'+str(i)+'-5.5'
            ws2['D'+ str(i)].number_format = '0'

        wb.save(dest)
    
#################################################
#################### FUNKCIE ####################
#################################################

    # Funkcia vytvori projektovu zlozku
    def createDir(self):
        self.path = self.projPath_value.get() 
        #self.tools = self.get_tool_lst()
        makedirs(self.path)
        # ! Zapracovat algoritmus pre uz existujuci priecinok
    
    # Funkcia vygeneruje oznacenie naradia
    def get_numbering(self):
        self.d_nad = self.prumNad_value.get()
        self.h_nad = self.vysNad_value.get()
        self.tlak = self.tlak_value.get()
        self.tvar_ramena = self.tvarRam_value.get()
        return(f"NA-{self.d_nad}-{self.h_nad}-{self.tlak}-{self.tvar_ramena}-")

    def get_init_val(self):
        self.toolsToCopy = self.get_tool_lst()

    def set_numbering(self):
        self.oznNar.delete(0,END)
        self.oznNar.insert(END,self.get_numbering())

    def get_reduction(self):
        # Priemerna redukcia
        self.dRam = float(self.prumRam.get())
        self.dKom = float(self.prumKom.get())
        self.tKom = float(self.tlKom.get())
        self.tVonkLak = float(self.tlIntLak.get())
        self.tVnutLak = float(self.tlExtLak.get())
        self.tLak = self.tVonkLak + self.tVnutLak
        self.n = int(self.pocTah.get())
        self.prum_red = round((((self.dRam - (self.dKom + 2*self.tKom + 2*self.tLak))/(self.n))/((self.dRam + self.dKom)/2) * 100),2)
        return self.prum_red

    def set_red(self):
        self.prumRed.delete(0,END)
        self.prumRed_value.set(self.get_reduction())

    # Funkcia na volbu naradia, ktore chceme navrhnut
    # Vrati slovnik "naradie": start. cislo oznacenia
    def get_tool_lst(self):
        self.toolsList = ["Stahovaci krouzky", 
                        "Chytaky", 
                        "Vodici pouzdra", 
                        "Navadeci krouzky", 
                        "Drzaky chytaku",
                        "Sroubove cepy",
                        "Trny",
                        "Pruziny"]
                    
        self.checkedTools = [self.stKr.get(),
                            self.chytaky.get(),
                            self.vodPzdra.get(),
                            self.navKr.get(),
                            self.drzChyt.get(),
                            self.sroubCepy.get(),
                            self.trny.get(),
                            self.pruziny.get()]

        self.initVals = [self.stKrInitVal.get(),
                        self.chytakyInitVal.get(),
                        self.vodPzdraInitVal.get(),
                        self.navKrInitVal.get(),
                        self.drzChytInitVal.get(),
                        self.sroubCepyInitVal.get(),
                        self.trnyInitVal.get(),
                        self.pruzinyInitVal.get()]
                        
        self.toolsRsm = list(zip(self.toolsList, self.initVals, self.checkedTools))
        self.designedTools = {tools[0]:tools[1] for tools in self.toolsRsm if tools[2]}
        #vraciam slovnik napr "chytaky":201
        return self.designedTools 

    # Funkcia na vypis suborov v zlozke pre ignore do copytree
    def ig_f(self, dir, files):
        return [f for f in files if os.path.isfile(os.path.join(dir,f))]

    # Metoda na ziskanie nazvu suboru bez pripony
    def getFileName(self, file):
        tempName = file.partition(".")[0]
        baseName = str(self.get_numbering())
        fileName = str(baseName.replace(tempName, baseName))
        return fileName

    # Metoda na ziskanie pripony suboru
    def getExt(self, file):
        return ("."+file.partition(".")[2])

    def copy_templates(self):
        self.toolsToCopy = self.get_tool_lst().keys()
        self.path = self.projPath_value.get()
        folders = walk(self.sabNar.get())
        foldLst = []
        for k in folders:
            foldLst.append(k[1])
        for i in self.toolsToCopy:
            copytree(self.sabNar.get() + "\\"+str(i), self.path+ "\\"+str(i), ignore=self.ig_f)  
        
    def copy_tools(self):
        self.toolsCount = int(self.pocTah.get())+1 #ziskam pocet tahov
        folders = listdir(self.path)        #ziskam strukturu projektovej zlozky
        toolsInitVal = list(self.get_tool_lst().values())
        self.temPath = self.sabNar.get()
 
        for i in folders:   # iterujem projektovou zlozkou
            if i in self.toolsToCopy:    # kontrola, ci je zlozka zo sablony
                files = listdir(self.temPath + "\\"+str(i))
                for j in files:
                    for k in range(1, self.toolsCount):
                        fileName = self.getFileName(str(j))
                        fileExt = self.getExt(str(j))
                        initValPos = list(self.toolsToCopy).index(i)
                        copy2(self.temPath + "\\"+str(i)+"\\"+str(j), self.path + "\\"+str(i)+"\\"+fileName+(str(toolsInitVal[initValPos]+k-1))+fileExt)

#################################################
####### KOPIROVANIE PARAMETROV DO EXCELOV #######
#################################################

# pre kazdy typ naradia:

# 1. iterujem v zlozke kazdy subor s priponou .xls
# 2. ulozim si nazov suboru do stringu
# 3. vyhladam string v hlavnom exceli
# 4. presmerujem sa na riadok, obsahujuci string nazvu suboru
# 5. urcim poziciu bunky
# 6. vytiahnem hodnotu z bunky
# 7. otvorim subor so stringom nazvu
# 8. vyhladam poziciu bunky, ktoru chcem prepisat
# 9. vlozim hodnotu

    def get_xcels(self):
        # vytvorit zoznam nazvov xls suborov v adresari konkretneho naradia BEZ PRIPONY
        
        # najst podla kazdeho stringu(suboru) v tom zozname ZHODNY riadok v hlavnom exceli navrh_naradi (stlpec B)
        # vytvorit pole hodnot daneho parametru (napr priemery chytaku D)
        # prekopirovat hodnoty z z pola hodnot rozmeru natvrdo do prislusnej bunky v exceli naradia (napr C26 -> B2)
        pass

#################################################
############# SPUSTENIE HLAVNEHO OKNA ###########
#################################################
    def exec(self):
        #self.createDir()
        self.copy_templates()
        self.create_XLS()
        self.copy_tools()
        
window=Tk()
Window(window)
window.mainloop()