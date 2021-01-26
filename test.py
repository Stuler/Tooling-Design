from os import mkdir
from re import A
import openpyxl
from openpyxl import Workbook, load_workbook
import openpyxl.utils
import openpyxl.cell.cell

'''lst1=["Stahovaci krouzky", 
    "Chytaky", 
    "Vodici pouzdra", 
    "Navadeci krouzky", 
    "Drzaky chytaku",
    "Sroubove cepy",
    "Trny",
    "Pruziny"]

lst2=[0,0,1,0,0,1,1,0]

zippedLst = list(zip(lst1,lst2))

validLst = [i[0] for i in zippedLst if i[1]==1]
print(validLst)

def createDir():
    path = ("C:\\Python\\Test") 
    
    mkdir(path + "\\")
    return path

print(createDir())'''

wb = Workbook()
ws1 = wb.active
ws1.title = "Data"

cells1 = openpyxl.utils.cell.rows_from_range(("A10:A20"))

for i in cells1:
    print(i[0])

cells2 = openpyxl.utils.cell.coordinate_to_tuple("A5")
ws1.append(cells2)

ws1['A4'] = 2

'''cellX = ws1['A4'].offset(row=19, column=0)
cellX.value = "4"'''

'''colLet = openpyxl.utils.cell.get_column_letter(4)
print(colLet)'''

'''ranBound = openpyxl.utils.cell.range_boundaries("D4:D20")
print(ranBound)'''

cellRange = 4 + 19
minCell = 4
maxCell = cellRange

col = openpyxl.cell.cell.get_column_letter(4)
ws1[''+ str(col) + str(minCell)] = "4"


    


wb.save('C:\\test.xlsx')